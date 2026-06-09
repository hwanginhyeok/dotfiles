#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
한국 개인사업자 세무 분석 결과를 엑셀(.xlsx) + HTML 리포트로 출력하는 생성기.

사용법:
    python3 gen_tax_report.py --input <json경로> [--outdir <출력디렉토리>]
    cat data.json | python3 gen_tax_report.py --outdir /tmp/out

출력:
    {사업자명}_세무분석_{귀속연도}_{작성일}.xlsx
    {사업자명}_세무리포트_{귀속연도}_{작성일}.html
"""

import argparse
import json
import math
import os
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# 공통 상수 / 색상 팔레트
# ─────────────────────────────────────────────
# 헤더 배경: 네이비
COLOR_HEADER_BG = "1A3C6E"
# 부제목 배경: 연파랑
COLOR_SUBHEADER_BG = "D6E4F7"
# 합계 행 배경: 연노랑
COLOR_TOTAL_BG = "FFF3CD"
# 강조(절감액) 배경: 연빨강
COLOR_ACCENT_BG = "FFE4E4"
# 장부 블록 배경
COLOR_BOOK_BG = "E8F4FD"
# 추계 블록 배경
COLOR_EST_BG = "FEF3E2"

FONT_NAME = "맑은 고딕"  # 없으면 openpyxl이 기본 폰트로 대체


# ─────────────────────────────────────────────
# 유틸 함수
# ─────────────────────────────────────────────
def safe_get(obj: dict, *keys, default=None):
    """중첩 dict에서 키를 순서대로 따라가며 값 반환. 없으면 default."""
    cur = obj
    for k in keys:
        if not isinstance(cur, dict):
            return default
        cur = cur.get(k, None)
        if cur is None:
            return default
    return cur


def sanitize_filename(name: str) -> str:
    """파일명에 쓸 수 없는 문자(공백, 특수문자)를 _로 치환."""
    return re.sub(r'[^\w가-힣\-]', '_', name)


def fmt_krw(value) -> str:
    """정수 금액을 한국 원화 형식으로 포맷 (예: 8,324,360원)."""
    try:
        return f"{int(value):,}원"
    except (TypeError, ValueError):
        return str(value)


def make_border(style="thin") -> Border:
    """모든 방향 테두리 생성."""
    side = Side(style=style)
    return Border(left=side, right=side, top=side, bottom=side)


def make_font(bold=False, color="000000", size=11) -> Font:
    """폰트 생성."""
    return Font(name=FONT_NAME, bold=bold, color=color, size=size)


def make_fill(hex_color: str) -> PatternFill:
    """단색 채우기 생성."""
    return PatternFill("solid", fgColor=hex_color)


def set_col_width(ws, col_idx: int, width: float):
    """열 너비 설정."""
    ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_section_title(ws, row: int, col_start: int, col_end: int, text: str):
    """
    섹션 제목 셀: 병합 + 굵게 + 네이비 배경 + 흰글씨 + 가운데 정렬.
    반환: 다음 행 번호.
    """
    ws.merge_cells(
        start_row=row, start_column=col_start,
        end_row=row, end_column=col_end
    )
    cell = ws.cell(row=row, column=col_start, value=text)
    cell.font = make_font(bold=True, color="FFFFFF", size=12)
    cell.fill = make_fill(COLOR_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = make_border()
    ws.row_dimensions[row].height = 22
    return row + 1


def write_table_header(ws, row: int, headers: list[tuple[int, str]]):
    """
    표 헤더 행: 열 인덱스·텍스트 튜플 리스트 받아서 연파랑 배경 + 굵게 + 가운데.
    반환: 다음 행 번호.
    """
    for col_idx, text in headers:
        cell = ws.cell(row=row, column=col_idx, value=text)
        cell.font = make_font(bold=True, color="1A3C6E", size=10)
        cell.fill = make_fill(COLOR_SUBHEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = make_border()
    ws.row_dimensions[row].height = 18
    return row + 1


def apply_money_format(cell):
    """금액 셀: 서식 + 우측 정렬."""
    cell.number_format = '#,##0"원"'
    cell.alignment = Alignment(horizontal="right", vertical="center")


def write_data_row(
    ws, row: int,
    values: list,  # (col_idx, value, is_money) 튜플 리스트
    fill_hex: str = None,
    bold: bool = False,
):
    """
    데이터 행 작성. is_money=True이면 금액 서식 적용.
    반환: 다음 행 번호.
    """
    for col_idx, value, is_money in values:
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.border = make_border()
        cell.font = make_font(bold=bold, size=10)
        if fill_hex:
            cell.fill = make_fill(fill_hex)
        if is_money and isinstance(value, (int, float)):
            apply_money_format(cell)
        else:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 17
    return row + 1


# ─────────────────────────────────────────────
# 시트1: "입력" 생성
# ─────────────────────────────────────────────
def build_sheet_input(ws, data: dict):
    """
    시트1: meta 정보 + 수입/비용/공제 표.
    열: A=항목, B=금액, C=증빙(비용 전용)
    """
    ws.title = "입력"

    # ── 열 너비 ──
    set_col_width(ws, 1, 36)  # A: 항목
    set_col_width(ws, 2, 20)  # B: 금액
    set_col_width(ws, 3, 20)  # C: 증빙

    row = 1

    # ── META 제목 ──
    row = write_section_title(ws, row, 1, 3, "📋 기본 정보")

    meta = data.get("meta", {})
    meta_fields = [
        ("사업자명", meta.get("사업자명", "")),
        ("업종", meta.get("업종", "")),
        ("업종코드", meta.get("업종코드", "")),
        ("과세유형", meta.get("과세유형", "")),
        ("귀속연도", meta.get("귀속연도", "")),
        ("작성일", meta.get("작성일", "")),
    ]
    for label, val in meta_fields:
        row = write_data_row(ws, row, [(1, label, False), (2, val, False), (3, "", False)])

    row += 1  # 빈 행

    # ── 수입 표 ──
    row = write_section_title(ws, row, 1, 3, "💰 수입 내역")
    row = write_table_header(ws, row, [(1, "항목"), (2, "금액"), (3, "비고")])

    income_list = data.get("수입", [])
    total_income = 0
    for item in income_list:
        amt = item.get("금액", 0) or 0
        total_income += amt
        row = write_data_row(ws, row, [
            (1, item.get("항목", ""), False),
            (2, amt, True),
            (3, item.get("비고", ""), False),
        ])
    # 합계 행
    row = write_data_row(ws, row, [
        (1, "합계", False),
        (2, total_income, True),
        (3, "", False),
    ], fill_hex=COLOR_TOTAL_BG, bold=True)

    row += 1  # 빈 행

    # ── 비용 표 ──
    row = write_section_title(ws, row, 1, 3, "🧾 비용 내역")
    row = write_table_header(ws, row, [(1, "항목"), (2, "금액"), (3, "증빙")])

    cost_list = data.get("비용", [])
    total_cost = 0
    for item in cost_list:
        amt = item.get("금액", 0) or 0
        total_cost += amt
        row = write_data_row(ws, row, [
            (1, item.get("항목", ""), False),
            (2, amt, True),
            (3, item.get("증빙", ""), False),
        ])
    # 합계 행
    row = write_data_row(ws, row, [
        (1, "합계", False),
        (2, total_cost, True),
        (3, "", False),
    ], fill_hex=COLOR_TOTAL_BG, bold=True)

    row += 1  # 빈 행

    # ── 공제 표 ──
    row = write_section_title(ws, row, 1, 2, "💳 소득공제 항목")
    # C열 병합 표시 위해 빈 셀 테두리만
    ws.cell(row=row - 1, column=3).border = make_border()

    row = write_table_header(ws, row, [(1, "항목"), (2, "금액"), (3, "")])

    deduction_list = data.get("공제", [])
    total_deduction = 0
    for item in deduction_list:
        amt = item.get("금액", 0) or 0
        total_deduction += amt
        row = write_data_row(ws, row, [
            (1, item.get("항목", ""), False),
            (2, amt, True),
            (3, "", False),
        ])
    # 합계 행
    row = write_data_row(ws, row, [
        (1, "합계", False),
        (2, total_deduction, True),
        (3, "", False),
    ], fill_hex=COLOR_TOTAL_BG, bold=True)


# ─────────────────────────────────────────────
# 시트2: "세액추정" 생성
# ─────────────────────────────────────────────
def build_sheet_tax_estimate(ws, data: dict):
    """
    시트2: 장부 vs 추계 나란히 비교 표 + 부가세.
    열 레이아웃:
        A: 항목명 | B: 장부 금액 | C: 추계 금액
    """
    ws.title = "세액추정"

    set_col_width(ws, 1, 28)  # A: 항목
    set_col_width(ws, 2, 22)  # B: 장부
    set_col_width(ws, 3, 22)  # C: 추계

    row = 1
    세액 = data.get("세액추정", {})
    장부 = 세액.get("장부", {})
    추계 = 세액.get("추계", {})

    # ── 비교 표 제목 ──
    row = write_section_title(ws, row, 1, 3, "📊 세액 추정 — 장부 vs 추계 비교")

    # 표 헤더
    row = write_table_header(ws, row, [
        (1, "항목"),
        (2, f"[장부] {장부.get('방식', '')}"),
        (3, f"[추계] {추계.get('방식', '')}"),
    ])

    # 비교 행 정의: (표시명, json키, is_money)
    compare_rows = [
        ("소득금액", "소득금액", True),
        ("소득공제", "소득공제", True),
        ("과세표준", "과세표준", True),
        ("산출세액", "산출세액", True),
        ("세액공제", "세액공제", True),
        ("결정세액", "결정세액", True),
        ("지방소득세", "지방소득세", True),
        ("총 세금 부담", "총부담", True),
        ("비고", "비고", False),
    ]

    for label, key, is_money in compare_rows:
        v_book = 장부.get(key, "")
        v_est = 추계.get(key, "")
        is_total = label == "총 세금 부담"

        row = write_data_row(ws, row, [
            (1, label, False),
            (2, v_book if v_book != "" else 0 if is_money else "", is_money),
            (3, v_est if v_est != "" else 0 if is_money else "", is_money),
        ], fill_hex=COLOR_TOTAL_BG if is_total else None, bold=is_total)

    row += 1  # 빈 행

    # ── 유리한 방식 + 절감액 강조 블록 ──
    유리 = 세액.get("유리한방식", "")
    절감 = 세액.get("절감액", 0) or 0

    row = write_section_title(ws, row, 1, 3, f"✅ 유리한 방식: {유리}   |   절감액: {절감:,}원")

    # 강조 셀 서식 재지정
    title_cell = ws.cell(row=row - 1, column=1)
    title_cell.fill = make_fill(COLOR_ACCENT_BG)
    title_cell.font = make_font(bold=True, color="C0392B", size=13)

    row += 1  # 빈 행

    # ── 부가세 표 ──
    row = write_section_title(ws, row, 1, 3, "🧾 부가세 (연간)")
    row = write_table_header(ws, row, [(1, "항목"), (2, "금액"), (3, "비고")])

    부가세 = 세액.get("부가세", {})
    vat_rows = [
        ("매출세액", 부가세.get("매출세액", 0), ""),
        ("매입세액(공제)", 부가세.get("매입세액", 0), ""),
        ("납부세액", 부가세.get("납부", 0), 부가세.get("비고", "")),
    ]
    for label, amt, note in vat_rows:
        is_total = label == "납부세액"
        row = write_data_row(ws, row, [
            (1, label, False),
            (2, amt, True),
            (3, note, False),
        ], fill_hex=COLOR_TOTAL_BG if is_total else None, bold=is_total)


# ─────────────────────────────────────────────
# 시트3: "절세전략" 생성
# ─────────────────────────────────────────────
def build_sheet_tax_tips(ws, data: dict):
    """
    시트3: 절세 전략(올해/내년) + 다음 액션 + 주의.
    열: A=제안/액션, B=효과, C=조건
    """
    ws.title = "절세전략"

    set_col_width(ws, 1, 38)  # A
    set_col_width(ws, 2, 22)  # B
    set_col_width(ws, 3, 28)  # C

    row = 1

    # ── 올해 절세 ──
    row = write_section_title(ws, row, 1, 3, "ⓐ 올해 바로 반영할 절세 포인트")
    row = write_table_header(ws, row, [(1, "제안"), (2, "효과"), (3, "조건")])

    for tip in data.get("절세_올해", []):
        row = write_data_row(ws, row, [
            (1, tip.get("제안", ""), False),
            (2, tip.get("효과", ""), False),
            (3, tip.get("조건", ""), False),
        ])

    row += 1  # 빈 행

    # ── 내년 세팅 ──
    row = write_section_title(ws, row, 1, 3, "ⓑ 내년을 위한 세팅")
    row = write_table_header(ws, row, [(1, "제안"), (2, "효과"), (3, "조건")])

    for tip in data.get("절세_내년", []):
        row = write_data_row(ws, row, [
            (1, tip.get("제안", ""), False),
            (2, tip.get("효과", ""), False),
            (3, tip.get("조건", ""), False),
        ])

    row += 1  # 빈 행

    # ── 다음 액션 ──
    row = write_section_title(ws, row, 1, 3, "📌 다음 액션")
    row = write_table_header(ws, row, [(1, "액션 항목"), (2, ""), (3, "")])

    for idx, action in enumerate(data.get("액션", []), start=1):
        cell1 = ws.cell(row=row, column=1, value=f"{idx}. {action}")
        cell1.border = make_border()
        cell1.font = make_font(size=10)
        cell1.alignment = Alignment(vertical="center", wrap_text=True)
        # B, C 병합
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        cell2 = ws.cell(row=row, column=2, value="")
        cell2.border = make_border()
        ws.row_dimensions[row].height = 20
        row += 1

    row += 1  # 빈 행

    # ── 주의 박스 ──
    row = write_section_title(ws, row, 1, 3, "⚠ 주의 사항")

    주의 = data.get("주의", "")
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 2, end_column=3)
    cell = ws.cell(row=row, column=1, value=주의)
    cell.font = make_font(size=10, color="7C4D00")
    cell.fill = make_fill("FFF8E1")
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border = make_border()
    ws.row_dimensions[row].height = 60


# ─────────────────────────────────────────────
# 엑셀 생성 메인
# ─────────────────────────────────────────────
def generate_xlsx(data: dict, outpath: Path) -> Path:
    """JSON 데이터로 엑셀 파일 생성. outpath를 반환."""
    wb = Workbook()

    # 기본 Sheet 삭제 후 시트 3개 추가
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    ws1 = wb.create_sheet("입력")
    ws2 = wb.create_sheet("세액추정")
    ws3 = wb.create_sheet("절세전략")

    build_sheet_input(ws1, data)
    build_sheet_tax_estimate(ws2, data)
    build_sheet_tax_tips(ws3, data)

    wb.save(outpath)
    return outpath


# ─────────────────────────────────────────────
# HTML 생성 메인
# ─────────────────────────────────────────────

# 내장 HTML 템플릿 (templates/report_template.html 없을 때 fallback)
_FALLBACK_HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<title>{사업자명} 세무 분석 리포트 ({귀속연도}귀속)</title>
<style>
  body {{ font-family: '맑은 고딕', system-ui, sans-serif; font-size: 13px; color: #1a1a2e; background: #f8f9fb; padding: 32px 16px; }}
  @media print {{ body {{ background: #fff; padding: 0; }} }}
  .report-wrap {{ max-width: 900px; margin: 0 auto; }}
  .report-header {{ background: linear-gradient(135deg, #1a3c6e 0%, #2563a8 100%); color: #fff; border-radius: 10px; padding: 28px 32px 22px; margin-bottom: 24px; }}
  .report-header h1 {{ font-size: 22px; font-weight: 700; margin-bottom: 10px; }}
  .meta-grid {{ display: grid; grid-template-columns: repeat(5, auto); gap: 8px 24px; margin-top: 12px; }}
  .meta-item {{ font-size: 12px; opacity: 0.88; }}
  .meta-item strong {{ display: block; font-size: 10px; opacity: 0.7; }}
  .disclaimer {{ margin-top: 14px; padding: 8px 12px; background: rgba(255,255,255,0.12); border-radius: 6px; font-size: 11px; opacity: 0.85; }}
  .card {{ background: #fff; border-radius: 8px; box-shadow: 0 1px 4px rgba(0,0,0,0.08); padding: 20px 24px; margin-bottom: 18px; }}
  .card-title {{ font-size: 14px; font-weight: 700; color: #1a3c6e; border-left: 4px solid #2563a8; padding-left: 10px; margin-bottom: 14px; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 12.5px; }}
  th {{ background: #1a3c6e; color: #fff; font-weight: 600; padding: 7px 10px; text-align: center; }}
  td {{ padding: 6px 10px; border-bottom: 1px solid #e8ecf2; }}
  tr:nth-child(even) td {{ background: #f5f7fb; }}
  td.num {{ text-align: right; }}
  .dday-urgent {{ color: #d32f2f; font-weight: 700; }}
  .winner-box {{ display: flex; align-items: center; gap: 16px; background: #e8f4fd; border: 2px solid #2563a8; border-radius: 8px; padding: 12px 18px; margin-bottom: 14px; }}
  .winner-value {{ font-size: 18px; font-weight: 800; color: #1a3c6e; }}
  .saving-value {{ font-size: 18px; font-weight: 800; color: #d32f2f; }}
  .winner-label {{ font-size: 12px; color: #4a5568; }}
  .tax-compare-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 14px; margin-bottom: 16px; }}
  .tax-col {{ border: 1px solid #e2e8f0; border-radius: 7px; overflow: hidden; }}
  .tax-col-header {{ padding: 8px 14px; font-weight: 700; font-size: 13px; text-align: center; }}
  .tax-col-header.book {{ background: #e8f4fd; color: #1a3c6e; }}
  .tax-col-header.estimate {{ background: #fef3e2; color: #7c4d00; }}
  .tax-col td {{ padding: 5px 12px; }}
  .tax-col tr.total-row td {{ font-weight: 700; background: #f0f4ff; }}
  .vat-inline {{ display: flex; gap: 10px; flex-wrap: wrap; }}
  .vat-chip {{ background: #f0f4ff; border: 1px solid #c7d7f5; border-radius: 5px; padding: 4px 12px; font-size: 12px; }}
  .tips-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 14px; }}
  .tip-card {{ border: 1px solid #e2e8f0; border-radius: 7px; padding: 12px 14px; }}
  .tip-card h4 {{ font-size: 12.5px; font-weight: 700; color: #1a3c6e; margin-bottom: 6px; }}
  .tip-row {{ font-size: 11.5px; color: #555; margin-bottom: 3px; }}
  .tip-effect {{ color: #2a7d2e; font-weight: 600; }}
  .action-list {{ list-style: none; padding: 0; }}
  .action-list li {{ display: flex; align-items: flex-start; gap: 10px; padding: 8px 0; border-bottom: 1px dashed #e2e8f0; font-size: 12.5px; }}
  .action-num {{ width: 22px; height: 22px; background: #2563a8; color: #fff; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-size: 11px; font-weight: 700; flex-shrink: 0; }}
  .warning-box {{ background: #fff8e1; border: 1px solid #ffcc02; border-left: 4px solid #f59e0b; border-radius: 6px; padding: 12px 16px; font-size: 12px; color: #6b4c00; margin-top: 16px; }}
  .profile-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 8px 20px; }}
  .profile-row {{ display: flex; gap: 8px; font-size: 12px; padding: 4px 0; border-bottom: 1px solid #eef1f6; }}
  .profile-key {{ color: #4a5568; flex: 0 0 80px; font-weight: 600; }}
  .section-num {{ display: inline-flex; align-items: center; justify-content: center; width: 20px; height: 20px; background: #2563a8; color: #fff; border-radius: 50%; font-size: 10px; font-weight: 700; margin-right: 6px; vertical-align: middle; }}
</style>
</head>
<body>
<div class="report-wrap">
  <div class="report-header">
    <h1>📊 {사업자명} &nbsp;|&nbsp; 세무 분석 리포트</h1>
    <div class="meta-grid">
      <div class="meta-item"><strong>업종</strong>{업종}</div>
      <div class="meta-item"><strong>업종코드</strong>{업종코드}</div>
      <div class="meta-item"><strong>과세유형</strong>{과세유형}</div>
      <div class="meta-item"><strong>귀속연도</strong>{귀속연도}년</div>
      <div class="meta-item"><strong>작성일</strong>{작성일}</div>
    </div>
    <div class="disclaimer">⚠ 본 리포트는 준비·시뮬레이션 보조용입니다. 최종 신고는 홈택스 또는 세무대리인에게 확정하세요.</div>
  </div>
  <div class="card">
    <div class="card-title"><span class="section-num">②</span>세무 프로파일</div>
    <div class="profile-grid">
      <div class="profile-row"><span class="profile-key">업종그룹</span><span>{프로파일_업종그룹}</span></div>
      <div class="profile-row"><span class="profile-key">장부의무</span><span>{프로파일_장부의무}</span></div>
      <div class="profile-row" style="grid-column:1/-1"><span class="profile-key">의무근거</span><span>{프로파일_장부의무근거}</span></div>
      <div class="profile-row"><span class="profile-key">경비율등급</span><span>{프로파일_경비율등급}</span></div>
      <div class="profile-row"><span class="profile-key">성실신고</span><span>{프로파일_성실신고}</span></div>
      <div class="profile-row" style="grid-column:1/-1"><span class="profile-key">복식부기전환</span><span>{프로파일_복식부기전환}</span></div>
    </div>
  </div>
  <div class="card">
    <div class="card-title"><span class="section-num">③</span>신고 일정</div>
    <table><thead><tr><th>마감일</th><th>신고 항목</th><th>D-day</th></tr></thead>
    <tbody>{일정_행}</tbody></table>
  </div>
  <div class="card">
    <div class="card-title"><span class="section-num">④</span>세액 추정 — 장부 vs 추계 비교</div>
    <div class="winner-box">
      <div><div class="winner-label">유리한 신고 방식</div><div class="winner-value">✅ {유리한방식}</div></div>
      <div style="border-left:2px solid #2563a8;padding-left:16px;"><div class="winner-label">절세 효과</div><div class="saving-value">🔻 절감액 {절감액}원</div></div>
    </div>
    <div class="tax-compare-grid">
      <div class="tax-col"><div class="tax-col-header book">📗 {장부_방식}</div><table>{장부_행}</table></div>
      <div class="tax-col"><div class="tax-col-header estimate">📙 {추계_방식}</div><table>{추계_행}</table></div>
    </div>
    <div style="margin-top:8px;"><div style="font-size:12px;font-weight:700;color:#555;margin-bottom:6px;">부가세 (연간)</div>
    <div class="vat-inline">{부가세_칩}</div></div>
  </div>
  <div class="card">
    <div class="card-title"><span class="section-num">⑤</span>절세 전략</div>
    <div style="font-size:12px;font-weight:700;color:#2a7d2e;margin-bottom:8px;">ⓐ 올해 바로 반영</div>
    <div class="tips-grid" style="margin-bottom:14px;">{절세_올해_카드}</div>
    <div style="font-size:12px;font-weight:700;color:#7c4d00;margin-bottom:8px;">ⓑ 내년 세팅</div>
    <div class="tips-grid">{절세_내년_카드}</div>
  </div>
  <div class="card">
    <div class="card-title"><span class="section-num">⑥</span>다음 액션</div>
    <ul class="action-list">{액션_목록}</ul>
    <div class="warning-box"><strong>⚠ 주의 </strong>{주의}</div>
  </div>
</div>
</body>
</html>"""


def build_html_일정_행(일정_list: list) -> str:
    """신고 일정 표 <tbody> 행 HTML 생성."""
    rows = []
    for item in 일정_list:
        dday = item.get("dday", "")
        # D-0은 빨간색 강조
        dday_cls = ' class="dday-urgent"' if "D-0" in dday else ""
        rows.append(
            f'<tr><td>{item.get("마감", "")}</td>'
            f'<td>{item.get("신고", "")}</td>'
            f'<td{dday_cls}>{dday}</td></tr>'
        )
    return "\n".join(rows)


def build_html_세액_행(세액_dict: dict) -> str:
    """장부 또는 추계 세액 비교 표 행 HTML 생성."""
    fields = [
        ("소득금액", "소득금액"),
        ("소득공제", "소득공제"),
        ("과세표준", "과세표준"),
        ("산출세액", "산출세액"),
        ("세액공제", "세액공제"),
        ("결정세액", "결정세액"),
        ("지방소득세", "지방소득세"),
        ("총 세금 부담", "총부담"),
        ("비고", "비고"),
    ]
    rows = []
    for label, key in fields:
        val = 세액_dict.get(key, "")
        is_total = key == "총부담"
        row_cls = ' class="total-row"' if is_total else ""
        if isinstance(val, (int, float)) and key != "비고":
            val_str = f"{int(val):,}원"
            td = f'<td style="text-align:right">{val_str}</td>'
        else:
            td = f"<td>{val}</td>"
        rows.append(f"<tr{row_cls}><td>{label}</td>{td}</tr>")
    return "\n".join(rows)


def build_html_부가세_칩(부가세: dict) -> str:
    """부가세 칩 HTML 생성."""
    chips = [
        ("매출세액", 부가세.get("매출세액", 0)),
        ("매입세액", 부가세.get("매입세액", 0)),
        ("납부세액", 부가세.get("납부", 0)),
    ]
    result = []
    for label, amt in chips:
        try:
            amt_str = f"{int(amt):,}원"
        except (TypeError, ValueError):
            amt_str = str(amt)
        result.append(
            f'<div class="vat-chip"><strong>{label}</strong> {amt_str}</div>'
        )
    note = 부가세.get("비고", "")
    if note:
        result.append(f'<div class="vat-chip">{note}</div>')
    return "\n".join(result)


def build_html_tip_cards(tips: list) -> str:
    """절세 전략 카드 HTML 생성."""
    cards = []
    for tip in tips:
        제안 = tip.get("제안", "")
        효과 = tip.get("효과", "")
        조건 = tip.get("조건", "")
        cards.append(
            f'<div class="tip-card">'
            f'<h4>{제안}</h4>'
            f'<div class="tip-row tip-effect">💡 {효과}</div>'
            f'<div class="tip-row">📋 조건: {조건}</div>'
            f'</div>'
        )
    return "\n".join(cards)


def build_html_액션_목록(액션_list: list) -> str:
    """다음 액션 번호 목록 HTML 생성."""
    items = []
    for idx, action in enumerate(액션_list, start=1):
        items.append(
            f'<li><span class="action-num">{idx}</span><span>{action}</span></li>'
        )
    return "\n".join(items)


def html_escape(text: str) -> str:
    """HTML 특수문자 이스케이프."""
    if not isinstance(text, str):
        text = str(text)
    return (text
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;"))


def generate_html(data: dict, outpath: Path, template_path: Path = None) -> Path:
    """JSON 데이터로 HTML 리포트 생성. outpath를 반환."""
    # 템플릿 로드 (파일 있으면 읽기, 없으면 내장 fallback)
    if template_path and template_path.exists():
        template = template_path.read_text(encoding="utf-8")
    else:
        template = _FALLBACK_HTML_TEMPLATE

    meta = data.get("meta", {})
    프로파일 = data.get("프로파일", {})
    세액 = data.get("세액추정", {})
    장부 = 세액.get("장부", {})
    추계 = 세액.get("추계", {})
    부가세 = 세액.get("부가세", {})

    절감액 = 세액.get("절감액", 0) or 0

    # placeholder 치환 딕셔너리 구성
    replacements = {
        "사업자명": html_escape(meta.get("사업자명", "")),
        "업종": html_escape(meta.get("업종", "")),
        "업종코드": html_escape(str(meta.get("업종코드", ""))),
        "과세유형": html_escape(meta.get("과세유형", "")),
        "귀속연도": html_escape(str(meta.get("귀속연도", ""))),
        "작성일": html_escape(meta.get("작성일", "")),
        # 프로파일
        "프로파일_업종그룹": html_escape(프로파일.get("업종그룹", "")),
        "프로파일_장부의무": html_escape(프로파일.get("장부의무", "")),
        "프로파일_장부의무근거": html_escape(프로파일.get("장부의무근거", "")),
        "프로파일_경비율등급": html_escape(프로파일.get("경비율등급", "")),
        "프로파일_복식부기전환": html_escape(프로파일.get("복식부기전환", "")),
        "프로파일_성실신고": html_escape(프로파일.get("성실신고", "")),
        # 일정
        "일정_행": build_html_일정_행(data.get("일정", [])),
        # 세액 추정
        "유리한방식": html_escape(세액.get("유리한방식", "")),
        "절감액": f"{int(절감액):,}",
        "장부_방식": html_escape(장부.get("방식", "")),
        "추계_방식": html_escape(추계.get("방식", "")),
        "장부_행": build_html_세액_행(장부),
        "추계_행": build_html_세액_행(추계),
        "부가세_칩": build_html_부가세_칩(부가세),
        # 절세 전략
        "절세_올해_카드": build_html_tip_cards(data.get("절세_올해", [])),
        "절세_내년_카드": build_html_tip_cards(data.get("절세_내년", [])),
        # 액션 + 주의
        "액션_목록": build_html_액션_목록(data.get("액션", [])),
        "주의": html_escape(data.get("주의", "")),
    }

    # str.replace 순차 치환 — CSS 중괄호({...})와 format_map 충돌 회피.
    # {KEY} 패턴을 찾아 순서대로 직접 교체.
    html_content = template
    for key, value in replacements.items():
        html_content = html_content.replace("{" + key + "}", str(value))

    outpath.write_text(html_content, encoding="utf-8")
    return outpath


# ─────────────────────────────────────────────
# 파일명 생성
# ─────────────────────────────────────────────
def build_output_paths(data: dict, outdir: Path) -> tuple[Path, Path]:
    """
    출력 파일 경로 2개(xlsx, html) 생성.
    파일명 패턴: {사업자명}_세무분석_{귀속연도}_{작성일}.xlsx
    """
    meta = data.get("meta", {})
    사업자명 = sanitize_filename(meta.get("사업자명", "미지정"))
    귀속연도 = sanitize_filename(str(meta.get("귀속연도", "0000")))
    작성일 = sanitize_filename(meta.get("작성일", "00000000"))

    base = f"{사업자명}_세무분석_{귀속연도}_{작성일}"
    xlsx_path = outdir / f"{base}.xlsx"
    html_path = outdir / f"{사업자명}_세무리포트_{귀속연도}_{작성일}.html"
    return xlsx_path, html_path


# ─────────────────────────────────────────────
# 진입점
# ─────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="한국 개인사업자 세무 분석 결과를 엑셀+HTML 리포트로 출력"
    )
    parser.add_argument(
        "--input", "-i",
        type=str,
        default=None,
        help="입력 JSON 파일 경로 (생략 시 stdin에서 읽기)"
    )
    parser.add_argument(
        "--outdir", "-o",
        type=str,
        default=".",
        help="출력 디렉토리 (기본: 현재 디렉토리)"
    )
    args = parser.parse_args()

    # ── JSON 로드 ──
    try:
        if args.input:
            with open(args.input, encoding="utf-8") as f:
                data = json.load(f)
        else:
            data = json.load(sys.stdin)
    except FileNotFoundError as e:
        print(f"[오류] 입력 파일을 찾을 수 없습니다: {e}", file=sys.stderr)
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"[오류] JSON 파싱 실패: {e}", file=sys.stderr)
        sys.exit(1)

    # ── 출력 디렉토리 준비 ──
    outdir = Path(args.outdir).resolve()
    outdir.mkdir(parents=True, exist_ok=True)

    # ── 파일 경로 결정 ──
    xlsx_path, html_path = build_output_paths(data, outdir)

    # ── HTML 템플릿 경로 (이 스크립트 기준 ../templates/report_template.html) ──
    script_dir = Path(__file__).parent
    template_path = script_dir.parent / "templates" / "report_template.html"

    # ── 파일 생성 ──
    try:
        generate_xlsx(data, xlsx_path)
    except Exception as e:
        print(f"[오류] 엑셀 생성 실패: {e}", file=sys.stderr)
        sys.exit(1)

    try:
        generate_html(data, html_path, template_path)
    except Exception as e:
        print(f"[오류] HTML 생성 실패: {e}", file=sys.stderr)
        sys.exit(1)

    # ── stdout에 생성된 파일 경로 출력 ──
    print(str(xlsx_path))
    print(str(html_path))


if __name__ == "__main__":
    main()
